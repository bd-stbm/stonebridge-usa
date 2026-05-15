"""Generate combined holdings + priced workbooks for every trust under Dyne Family (US).

Pure offline analysis (Holdings + GWM) plus yfinance + OpenFIGI for pricing.
NO new Masttro API calls.

Trust selection: descendant of Dyne Family (US) (102_93356) whose alias OR name
contains "trust" AND has at least one direct-child investment account
(bankBroker AND accountNumber populated). Currently 10 trusts.

Outputs:
- reports/DyneFamilyUS_allTrusts_holdings_<YYYYMMDD-HHMMSS>.xlsx
- reports/DyneFamilyUS_allTrusts_priced_<YYYYMMDD-HHMMSS>.xlsx

Each workbook has a Summary sheet + one detail sheet per trust.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
import urllib.error
import urllib.request
import warnings
from collections import defaultdict
from pathlib import Path

import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESPONSES = PROJECT_ROOT / "responses"
REPORTS = PROJECT_ROOT / "reports"

DYNE_US_NODE = "102_93356"
OPENFIGI_URL = "https://api.openfigi.com/v3/mapping"
US_EXCH_PRIORITY = ["US", "UN", "UQ", "UA", "UR", "UF", "UV", "UB", "UW", "UP"]
COUNTRY_SUFFIX_RE = re.compile(r"_(US|AU|LN|GB|FR|DE|JP|CA|HK)$", re.IGNORECASE)


def normalize_ticker(t):
    if not t:
        return None
    t = t.strip().upper()
    t = COUNTRY_SUFFIX_RE.sub("", t)
    t = t.replace("/", "-")
    return t or None


def sheet_safe(name, taken):
    """Slug into a valid <=31-char sheet name, ensure uniqueness."""
    clean = re.sub(r"[\\/?*\[\]:]", "", name or "Trust")[:31].strip()
    if not clean:
        clean = "Trust"
    base = clean
    n = 2
    while clean in taken:
        suffix = f"_{n}"
        clean = (base[: 31 - len(suffix)] + suffix)
        n += 1
    taken.add(clean)
    return clean


def latest(pattern):
    return sorted(RESPONSES.glob(pattern))[-1]


def load(pattern):
    p = latest(pattern)
    print(f"loaded: {p.relative_to(PROJECT_ROOT)}  ({p.stat().st_size:,} bytes)")
    return json.loads(p.read_text(encoding="utf-8"))


# ----------------------------------------------------------------------------
# Load + index GWM
# ----------------------------------------------------------------------------
gwm = load("GWM-7693_*.json")
holdings = load("Holdings-7693_*.json")

nodes_by_id = {n["nodeId"]: n for n in gwm}
children_of = defaultdict(list)
for n in gwm:
    children_of[n.get("parentNodeId")].append(n["nodeId"])


def descendants(root):
    out = {root}
    stack = [root]
    while stack:
        nid = stack.pop()
        for c in children_of.get(nid, []):
            if c not in out:
                out.add(c)
                stack.append(c)
    return out


# ----------------------------------------------------------------------------
# Identify trusts
# ----------------------------------------------------------------------------
dyne_us_subtree = descendants(DYNE_US_NODE)

def is_trust_candidate(n):
    alias = (n.get("alias") or "").lower()
    name = (n.get("name") or "").lower()
    if "trust" not in alias and "trust" not in name:
        return False
    # Must have at least one direct-child investment account.
    for c in children_of.get(n["nodeId"], []):
        cn = nodes_by_id[c]
        if (cn.get("bankBroker") or "").strip() and (cn.get("accountNumber") or "").strip():
            return True
    return False


trust_nodes_raw = [
    nodes_by_id[nid] for nid in dyne_us_subtree
    if nid != DYNE_US_NODE and is_trust_candidate(nodes_by_id[nid])
]


def owner_of(nid):
    pid = nodes_by_id[nid].get("parentNodeId")
    if pid and pid in nodes_by_id:
        return nodes_by_id[pid].get("alias") or nodes_by_id[pid].get("name") or pid
    return ""


def account_number_fingerprint(trust_nid):
    """Frozenset of direct-child investment-account numbers. Trust nodes with
    identical fingerprints are the same legal trust seen via different
    beneficial-owner pathways and should be collapsed (full-duplication pattern,
    unlike SEP's fractional split)."""
    nums = set()
    for c in children_of.get(trust_nid, []):
        cn = nodes_by_id[c]
        bb = (cn.get("bankBroker") or "").strip()
        an = (cn.get("accountNumber") or "").strip()
        if bb and an:
            nums.add(an)
    return frozenset(nums)


# Dedupe: group by fingerprint, keep first node per group.
seen_fingerprints = {}
trust_nodes = []
duplicates = []
trust_nodes_raw.sort(key=lambda n: -(n.get("valuation") or 0))
for tn in trust_nodes_raw:
    fp = account_number_fingerprint(tn["nodeId"])
    if not fp:
        continue
    if fp in seen_fingerprints:
        duplicates.append((tn, seen_fingerprints[fp]))
    else:
        seen_fingerprints[fp] = tn
        trust_nodes.append(tn)

print(f"\nQualifying trusts under Dyne Family (US): {len(trust_nodes_raw)}")
print(f"Distinct legal trusts (after collapsing beneficial-owner duplicates): "
      f"{len(trust_nodes)}")
if duplicates:
    print(f"Collapsed duplicates (same account-number fingerprint as canonical):")
    for dup, canon in duplicates:
        dup_alias = dup.get("alias") or dup.get("name")
        canon_alias = canon.get("alias") or canon.get("name")
        print(f"  dropped {dup['nodeId']} ({dup_alias!r} under "
              f"{owner_of(dup['nodeId'])!r}) — same accounts as canonical "
              f"{canon['nodeId']} ({canon_alias!r} under {owner_of(canon['nodeId'])!r})")

print(f"\nCanonical trust list:")
for tn in trust_nodes:
    alias = tn.get("alias") or tn.get("name")
    print(f"  {tn['nodeId']:<14}  {alias[:48]:<50}  owner={owner_of(tn['nodeId'])[:20]:<22} "
          f"val={tn.get('valuation'):>14,.0f}")


# ----------------------------------------------------------------------------
# For each trust, find direct-child accounts and the Holdings rows under them
# ----------------------------------------------------------------------------
trust_data = []  # list of dicts: {trust, accounts, holdings, owner}
for tn in trust_nodes:
    accts = []
    for cid in children_of.get(tn["nodeId"], []):
        cn = nodes_by_id[cid]
        bb = (cn.get("bankBroker") or "").strip()
        an = (cn.get("accountNumber") or "").strip()
        if bb and an:
            accts.append(cn)
    acct_nid_set = {a["nodeId"] for a in accts}
    holds = [h for h in holdings if h.get("nodeId") in acct_nid_set]
    trust_data.append({
        "trust": tn,
        "owner": owner_of(tn["nodeId"]),
        "accounts": accts,
        "account_lookup": {a["nodeId"]: a for a in accts},
        "acct_nid_set": acct_nid_set,
        "holdings": holds,
    })

total_rows = sum(len(t["holdings"]) for t in trust_data)
total_mv = sum(
    (h.get("localMarketValue") or 0)
    for t in trust_data for h in t["holdings"]
)
print(f"\nTotal Holdings rows across all trusts: {total_rows}")
print(f"Total NAV (USD, sum of localMarketValue): {total_mv:,.2f}")


# ----------------------------------------------------------------------------
# Build the master ticker map (dedupe across all trusts)
# ----------------------------------------------------------------------------
# row_ticker[(trust_idx, row_idx)] = normalised ticker (or None)
row_ticker = {}
unique_tickers = set()
for t_i, td in enumerate(trust_data):
    for r_i, h in enumerate(td["holdings"]):
        nt = normalize_ticker(h.get("ticker"))
        if nt and h.get("assetClass") == "Equity":
            row_ticker[(t_i, r_i)] = nt
            unique_tickers.add(nt)

print(f"\nDistinct equity tickers across all trusts: {len(unique_tickers)}")


# ----------------------------------------------------------------------------
# yfinance batch (pass 1)
# ----------------------------------------------------------------------------
def fetch_prices(tickers, chunk=50):
    ok = {}
    failed = []
    for i in range(0, len(tickers), chunk):
        batch = tickers[i:i + chunk]
        print(f"  batch {i//chunk + 1}: {len(batch)} tickers...", end=" ", flush=True)
        try:
            data = yf.download(
                tickers=batch, period="5d", group_by="ticker",
                auto_adjust=False, progress=False, threads=True,
            )
        except Exception as e:
            print(f"ERR: {e}")
            failed.extend(batch)
            continue
        got = 0
        for t in batch:
            try:
                closes = data["Close"].dropna() if len(batch) == 1 else data[t]["Close"].dropna()
                if len(closes):
                    ok[t] = (float(closes.iloc[-1]), closes.index[-1].date())
                    got += 1
                else:
                    failed.append(t)
            except (KeyError, TypeError):
                failed.append(t)
        print(f"got {got}/{len(batch)}")
    return ok, failed


print(f"\nFetching prices from yfinance (pass 1)...")
prices, failures = fetch_prices(sorted(unique_tickers))
print(f"  Priced: {len(prices)}/{len(unique_tickers)}   Failed: {len(failures)}")


# ----------------------------------------------------------------------------
# OpenFIGI fallback
# ----------------------------------------------------------------------------
def openfigi_resolve(isins, chunk=10):
    out = {}
    if not isins:
        return out
    for i in range(0, len(isins), chunk):
        batch = isins[i:i + chunk]
        body = json.dumps(
            [{"idType": "ID_ISIN", "idValue": v} for v in batch]
        ).encode("utf-8")
        req = urllib.request.Request(
            OPENFIGI_URL, data=body,
            headers={"Content-Type": "application/json", "Accept": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                response = json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            print(f"  OpenFIGI batch error: {type(e).__name__}: {e}")
            for v in batch:
                out[v] = None
            continue
        for isin_value, result in zip(batch, response):
            matches = result.get("data") or []
            chosen = None
            for ex in US_EXCH_PRIORITY:
                chosen = next(
                    (m for m in matches if m.get("exchCode") == ex and m.get("ticker")),
                    None,
                )
                if chosen:
                    break
            if not chosen:
                chosen = next((m for m in matches if m.get("ticker")), None)
            out[isin_value] = chosen["ticker"] if chosen else None
    return out


# For each failed ticker, find any ISIN we can use.
failed_set = set(failures)
isin_by_failed_ticker = {}
for (t_i, r_i), tk in row_ticker.items():
    if tk in failed_set and tk not in isin_by_failed_ticker:
        h = trust_data[t_i]["holdings"][r_i]
        isin = (h.get("isin") or "").strip()
        if isin:
            isin_by_failed_ticker[tk] = isin
isins_to_resolve = sorted(set(isin_by_failed_ticker.values()))
print(f"\nOpenFIGI fallback: resolving {len(isins_to_resolve)} ISINs from "
      f"{len(failed_set)} failed tickers...")
isin_to_ticker = openfigi_resolve(isins_to_resolve)
resolved = sum(1 for v in isin_to_ticker.values() if v)
print(f"  Resolved: {resolved}/{len(isins_to_resolve)}")

# Build ticker remap (old → new via OpenFIGI), then yfinance pass 2.
ticker_remap = {}
for old, isin in isin_by_failed_ticker.items():
    new = isin_to_ticker.get(isin)
    if new:
        nn = normalize_ticker(new)
        if nn and nn != old:
            ticker_remap[old] = nn

new_unique = sorted(set(ticker_remap.values()) - set(prices.keys()))
print(f"  New unique tickers to try via yfinance: {len(new_unique)}")
if new_unique:
    print(f"\nFetching prices from yfinance (pass 2: OpenFIGI tickers)...")
    extra_prices, _ = fetch_prices(new_unique)
    print(f"  Priced: {len(extra_prices)}/{len(new_unique)}")
    prices.update(extra_prices)

# Track per-row ticker source.
ticker_source = {}
for key, tk in list(row_ticker.items()):
    if tk in prices:
        ticker_source[key] = "masttro"
    elif tk in ticker_remap and ticker_remap[tk] in prices:
        row_ticker[key] = ticker_remap[tk]
        ticker_source[key] = "openfigi"


# ----------------------------------------------------------------------------
# Workbook 1: Holdings (no pricing)
# ----------------------------------------------------------------------------
HOLDINGS_COLS = [
    ("Account Alias", "text"), ("Custodian", "text"), ("Account Number", "text"),
    ("Asset Name", "text"), ("ISIN", "text"), ("Ticker", "text"),
    ("SEDOL", "text"), ("CUSIP", "text"), ("Asset Class", "text"),
    ("Security Type", "text"), ("Sector", "text"), ("Geographic Exposure", "text"),
    ("Quantity", "qty"), ("Local CCY", "text"),
    ("Price (USD)", "price"), ("Market Value (USD)", "money"),
    ("Accrued Interest (USD)", "money"), ("Unit Cost (USD)", "price"),
    ("Total Cost (USD)", "money"),
    ("securityId", "text"), ("nodeId", "text"),
]

PRICED_COLS = [
    ("Account Alias", "text"), ("Custodian", "text"), ("Account Number", "text"),
    ("Asset Name", "text"), ("ISIN", "text"),
    ("Ticker (Masttro)", "text"), ("Ticker (yfinance)", "text"), ("Ticker Source", "text"),
    ("SEDOL", "text"), ("CUSIP", "text"), ("Asset Class", "text"),
    ("Security Type", "text"), ("Sector", "text"), ("Geographic Exposure", "text"),
    ("Quantity", "qty"), ("Local CCY", "text"),
    ("Masttro Price (USD)", "price"), ("yfinance Price (USD)", "price"),
    ("yfinance As-of", "text"),
    ("Δ Price (USD)", "price"), ("Δ Price %", "pct"),
    ("Masttro MV (USD)", "money"), ("Refreshed MV (USD)", "money"),
    ("Refreshed Δ MV (USD)", "money"),
    ("Accrued Interest (USD)", "money"), ("Unit Cost (USD)", "price"),
    ("Total Cost (USD)", "money"),
    ("securityId", "text"), ("nodeId", "text"),
]

money_fmt = '#,##0.00;(#,##0.00);"-"'
qty_fmt = '#,##0.0000;(#,##0.0000);"-"'
price_fmt = '#,##0.0000'
pct_fmt = '0.00%;-0.00%;"-"'

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True, italic=True)


def style_header(ws, cols):
    for c_idx, (name, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"


def num_format(cell, kind, v):
    if not isinstance(v, (int, float)):
        return
    if kind == "money":
        cell.number_format = money_fmt
    elif kind == "qty":
        cell.number_format = qty_fmt
    elif kind == "price":
        cell.number_format = price_fmt
    elif kind == "pct":
        cell.number_format = pct_fmt


def emit_subtotal(ws, cols, money_cols, r_idx, start_row, label):
    """Write a subtotal row using Excel SUM formulas. Returns new r_idx."""
    end_row = r_idx - 1
    if end_row < start_row:
        return r_idx
    for c_idx, (cname, _) in enumerate(cols, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = SUBTOTAL_FILL
        cell.font = SUBTOTAL_FONT
        if c_idx == 1:
            cell.value = label
        elif cname in money_cols:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            cell.number_format = money_fmt
    return r_idx + 1


# ---- Build holdings workbook
print(f"\nBuilding holdings workbook...")
wb_h = Workbook()
wb_h.remove(wb_h.active)
summary_h = wb_h.create_sheet(title="Summary")

summary_h_cols = [
    ("#", "int"), ("Trust", "text"), ("Owner", "text"), ("nodeId", "text"),
    ("# Accounts", "int"), ("# Rows", "int"), ("Total MV (USD)", "money"),
]
style_header(summary_h, summary_h_cols)

taken_sheet_names = {"Summary"}
holdings_money_cols = {"Market Value (USD)", "Accrued Interest (USD)", "Total Cost (USD)"}

for t_i, td in enumerate(trust_data):
    tn = td["trust"]
    alias = tn.get("alias") or tn.get("name") or tn["nodeId"]
    sheet_name = sheet_safe(alias, taken_sheet_names)
    ws = wb_h.create_sheet(title=sheet_name)
    style_header(ws, HOLDINGS_COLS)

    # Sort holdings by account valuation desc, then by MV desc within account.
    acct_val = {nid: a.get("valuation") or 0 for nid, a in td["account_lookup"].items()}
    sorted_hold = sorted(
        td["holdings"],
        key=lambda h: (-(acct_val.get(h["nodeId"], 0)), -(h.get("localMarketValue") or 0)),
    )

    r_idx = 2
    current_acct = None
    acct_start = 2

    trust_mv = 0.0
    for h in sorted_hold:
        acct = td["account_lookup"][h["nodeId"]]
        new_acct = acct.get("alias")
        if current_acct is None:
            current_acct = new_acct
            acct_start = r_idx
        elif new_acct != current_acct:
            r_idx = emit_subtotal(
                ws, HOLDINGS_COLS, holdings_money_cols,
                r_idx, acct_start, f"Subtotal: {current_acct}",
            )
            current_acct = new_acct
            acct_start = r_idx

        qty = h.get("quantity")
        if isinstance(qty, str):
            try:
                qty = float(qty)
            except ValueError:
                pass

        values = {
            "Account Alias": acct.get("alias"),
            "Custodian": acct.get("bankBroker"),
            "Account Number": acct.get("accountNumber"),
            "Asset Name": h.get("assetName"),
            "ISIN": h.get("isin"),
            "Ticker": h.get("ticker"),
            "SEDOL": h.get("sedol"),
            "CUSIP": h.get("cusip"),
            "Asset Class": h.get("assetClass"),
            "Security Type": h.get("securityType"),
            "Sector": h.get("sector"),
            "Geographic Exposure": h.get("geographicExposure"),
            "Quantity": qty,
            "Local CCY": h.get("localCCY"),
            "Price (USD)": h.get("price"),
            "Market Value (USD)": h.get("localMarketValue"),
            "Accrued Interest (USD)": h.get("localAccruedInterest"),
            "Unit Cost (USD)": h.get("unitCost"),
            "Total Cost (USD)": h.get("totalCost"),
            "securityId": h.get("securityId"),
            "nodeId": h.get("nodeId"),
        }
        for c_idx, (cname, kind) in enumerate(HOLDINGS_COLS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=values[cname])
            num_format(cell, kind, values[cname])
        if isinstance(values["Market Value (USD)"], (int, float)):
            trust_mv += values["Market Value (USD)"]
        r_idx += 1

    if current_acct is not None:
        r_idx = emit_subtotal(
            ws, HOLDINGS_COLS, holdings_money_cols,
            r_idx, acct_start, f"Subtotal: {current_acct}",
        )

    # Grand total for this trust
    for c_idx, (cname, _) in enumerate(HOLDINGS_COLS, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        if c_idx == 1:
            cell.value = "GRAND TOTAL"
        elif cname == "Market Value (USD)":
            cell.value = trust_mv
            cell.number_format = money_fmt

    # Column widths (simple, generous)
    for c_idx, (cname, _) in enumerate(HOLDINGS_COLS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(40, len(cname) + 2))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HOLDINGS_COLS))}1"

    # Summary row
    s_row = t_i + 2
    summary_values = [
        t_i + 1,
        alias,
        td["owner"],
        tn["nodeId"],
        len(td["accounts"]),
        len(sorted_hold),
        trust_mv,
    ]
    for c_idx, (sname, skind) in enumerate(summary_h_cols, start=1):
        cell = summary_h.cell(row=s_row, column=c_idx, value=summary_values[c_idx - 1])
        if skind == "money":
            cell.number_format = money_fmt
        elif skind == "int":
            cell.number_format = "0"

# Summary grand total row
grand_row = len(trust_data) + 2
for c_idx, (sname, skind) in enumerate(summary_h_cols, start=1):
    cell = summary_h.cell(row=grand_row, column=c_idx)
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    if c_idx == 1:
        cell.value = "TOTAL"
    elif sname == "# Accounts":
        cell.value = sum(len(t["accounts"]) for t in trust_data)
        cell.number_format = "0"
    elif sname == "# Rows":
        cell.value = sum(len(t["holdings"]) for t in trust_data)
        cell.number_format = "0"
    elif sname == "Total MV (USD)":
        # Sum local MV across all rows
        cell.value = sum(
            (h.get("localMarketValue") or 0)
            for t in trust_data for h in t["holdings"]
        )
        cell.number_format = money_fmt

for c_idx, (sname, _) in enumerate(summary_h_cols, start=1):
    summary_h.column_dimensions[get_column_letter(c_idx)].width = max(14, min(40, len(sname) + 4))

REPORTS.mkdir(exist_ok=True)
ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
holdings_out = REPORTS / f"DyneFamilyUS_allTrusts_holdings_{ts}.xlsx"
wb_h.save(holdings_out)
print(f"Wrote: {holdings_out.relative_to(PROJECT_ROOT)}")


# ---- Build priced workbook
print(f"\nBuilding priced workbook...")
wb_p = Workbook()
wb_p.remove(wb_p.active)
summary_p = wb_p.create_sheet(title="Summary")

summary_p_cols = [
    ("#", "int"), ("Trust", "text"), ("Owner", "text"), ("nodeId", "text"),
    ("# Accounts", "int"), ("# Rows", "int"), ("# Priced", "int"), ("% Priced", "pct"),
    ("Masttro MV (USD)", "money"),
    ("Refreshed MV (USD, priced rows)", "money"),
    ("Δ MV (USD)", "money"), ("Δ %", "pct"),
]
style_header(summary_p, summary_p_cols)

taken_sheet_names = {"Summary"}
priced_money_cols = {
    "Masttro MV (USD)", "Refreshed MV (USD)", "Refreshed Δ MV (USD)",
    "Accrued Interest (USD)", "Total Cost (USD)",
}

for t_i, td in enumerate(trust_data):
    tn = td["trust"]
    alias = tn.get("alias") or tn.get("name") or tn["nodeId"]
    sheet_name = sheet_safe(alias, taken_sheet_names)
    ws = wb_p.create_sheet(title=sheet_name)
    style_header(ws, PRICED_COLS)

    acct_val = {nid: a.get("valuation") or 0 for nid, a in td["account_lookup"].items()}
    sorted_idx = sorted(
        range(len(td["holdings"])),
        key=lambda i: (
            -(acct_val.get(td["holdings"][i]["nodeId"], 0)),
            -(td["holdings"][i].get("localMarketValue") or 0),
        ),
    )

    r_idx = 2
    current_acct = None
    acct_start = 2

    trust_masttro_mv = 0.0
    trust_refreshed_mv = 0.0
    trust_delta = 0.0
    trust_priced_count = 0

    for r_i in sorted_idx:
        h = td["holdings"][r_i]
        acct = td["account_lookup"][h["nodeId"]]
        new_acct = acct.get("alias")
        if current_acct is None:
            current_acct = new_acct
            acct_start = r_idx
        elif new_acct != current_acct:
            r_idx = emit_subtotal(
                ws, PRICED_COLS, priced_money_cols,
                r_idx, acct_start, f"Subtotal: {current_acct}",
            )
            current_acct = new_acct
            acct_start = r_idx

        yf_ticker = row_ticker.get((t_i, r_i))
        yf_price = None
        yf_date = None
        yf_source = None
        if yf_ticker and yf_ticker in prices:
            yf_price, yf_date = prices[yf_ticker]
            yf_source = ticker_source.get((t_i, r_i), "masttro")
            trust_priced_count += 1

        mp = h.get("price")
        masttro_usd_mv = h.get("localMarketValue") or 0
        delta_price = None
        delta_pct = None
        refreshed_usd_mv = None
        delta_usd_mv = None
        if (
            isinstance(mp, (int, float)) and isinstance(yf_price, (int, float))
            and mp != 0 and masttro_usd_mv
        ):
            delta_price = yf_price - mp
            delta_pct = delta_price / mp
            refreshed_usd_mv = masttro_usd_mv * (yf_price / mp)
            delta_usd_mv = refreshed_usd_mv - masttro_usd_mv

        qty = h.get("quantity")
        if isinstance(qty, str):
            try:
                qty = float(qty)
            except ValueError:
                pass

        values = {
            "Account Alias": acct.get("alias"),
            "Custodian": acct.get("bankBroker"),
            "Account Number": acct.get("accountNumber"),
            "Asset Name": h.get("assetName"),
            "ISIN": h.get("isin"),
            "Ticker (Masttro)": h.get("ticker"),
            "Ticker (yfinance)": yf_ticker if yf_price is not None else None,
            "Ticker Source": yf_source,
            "SEDOL": h.get("sedol"),
            "CUSIP": h.get("cusip"),
            "Asset Class": h.get("assetClass"),
            "Security Type": h.get("securityType"),
            "Sector": h.get("sector"),
            "Geographic Exposure": h.get("geographicExposure"),
            "Quantity": qty,
            "Local CCY": h.get("localCCY"),
            "Masttro Price (USD)": mp,
            "yfinance Price (USD)": yf_price,
            "yfinance As-of": yf_date.isoformat() if yf_date else None,
            "Δ Price (USD)": delta_price,
            "Δ Price %": delta_pct,
            "Masttro MV (USD)": masttro_usd_mv,
            "Refreshed MV (USD)": refreshed_usd_mv,
            "Refreshed Δ MV (USD)": delta_usd_mv,
            "Accrued Interest (USD)": h.get("localAccruedInterest"),
            "Unit Cost (USD)": h.get("unitCost"),
            "Total Cost (USD)": h.get("totalCost"),
            "securityId": h.get("securityId"),
            "nodeId": h.get("nodeId"),
        }
        for c_idx, (cname, kind) in enumerate(PRICED_COLS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=values[cname])
            num_format(cell, kind, values[cname])
        trust_masttro_mv += masttro_usd_mv
        if isinstance(refreshed_usd_mv, (int, float)):
            trust_refreshed_mv += refreshed_usd_mv
        if isinstance(delta_usd_mv, (int, float)):
            trust_delta += delta_usd_mv
        r_idx += 1

    if current_acct is not None:
        r_idx = emit_subtotal(
            ws, PRICED_COLS, priced_money_cols,
            r_idx, acct_start, f"Subtotal: {current_acct}",
        )

    # Grand total per trust
    for c_idx, (cname, _) in enumerate(PRICED_COLS, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        if c_idx == 1:
            cell.value = "GRAND TOTAL"
        elif cname == "Masttro MV (USD)":
            cell.value = trust_masttro_mv
            cell.number_format = money_fmt
        elif cname == "Refreshed MV (USD)":
            cell.value = trust_refreshed_mv
            cell.number_format = money_fmt
        elif cname == "Refreshed Δ MV (USD)":
            cell.value = trust_delta
            cell.number_format = money_fmt

    for c_idx, (cname, _) in enumerate(PRICED_COLS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(40, len(cname) + 2))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PRICED_COLS))}1"

    # Summary row
    pct_priced = (trust_priced_count / len(td["holdings"])) if td["holdings"] else 0
    delta_pct_total = (trust_delta / trust_masttro_mv) if trust_masttro_mv else 0
    s_row = t_i + 2
    s_values = [
        t_i + 1, alias, td["owner"], tn["nodeId"],
        len(td["accounts"]), len(td["holdings"]), trust_priced_count,
        pct_priced, trust_masttro_mv, trust_refreshed_mv,
        trust_delta, delta_pct_total,
    ]
    for c_idx, (sname, skind) in enumerate(summary_p_cols, start=1):
        cell = summary_p.cell(row=s_row, column=c_idx, value=s_values[c_idx - 1])
        if skind == "money":
            cell.number_format = money_fmt
        elif skind == "pct":
            cell.number_format = pct_fmt
        elif skind == "int":
            cell.number_format = "0"

# Summary grand total
grand_row = len(trust_data) + 2
totals = {
    "# Accounts": sum(len(t["accounts"]) for t in trust_data),
    "# Rows": sum(len(t["holdings"]) for t in trust_data),
}
total_masttro = sum(
    (h.get("localMarketValue") or 0)
    for t in trust_data for h in t["holdings"]
)
# Recompute refresh totals + priced count from the per-trust rows we just wrote
total_priced_count = 0
total_refreshed = 0.0
total_delta = 0.0
for t_i, td in enumerate(trust_data):
    for r_i, h in enumerate(td["holdings"]):
        yf_ticker = row_ticker.get((t_i, r_i))
        if yf_ticker and yf_ticker in prices:
            total_priced_count += 1
            mp = h.get("price")
            yf_price = prices[yf_ticker][0]
            masttro_usd_mv = h.get("localMarketValue") or 0
            if isinstance(mp, (int, float)) and mp != 0 and masttro_usd_mv:
                refreshed = masttro_usd_mv * (yf_price / mp)
                total_refreshed += refreshed
                total_delta += refreshed - masttro_usd_mv

for c_idx, (sname, skind) in enumerate(summary_p_cols, start=1):
    cell = summary_p.cell(row=grand_row, column=c_idx)
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    if c_idx == 1:
        cell.value = "TOTAL"
    elif sname == "# Accounts":
        cell.value = totals["# Accounts"]
        cell.number_format = "0"
    elif sname == "# Rows":
        cell.value = totals["# Rows"]
        cell.number_format = "0"
    elif sname == "# Priced":
        cell.value = total_priced_count
        cell.number_format = "0"
    elif sname == "% Priced":
        cell.value = (total_priced_count / totals["# Rows"]) if totals["# Rows"] else 0
        cell.number_format = pct_fmt
    elif sname == "Masttro MV (USD)":
        cell.value = total_masttro
        cell.number_format = money_fmt
    elif sname == "Refreshed MV (USD, priced rows)":
        cell.value = total_refreshed
        cell.number_format = money_fmt
    elif sname == "Δ MV (USD)":
        cell.value = total_delta
        cell.number_format = money_fmt
    elif sname == "Δ %":
        cell.value = (total_delta / total_masttro) if total_masttro else 0
        cell.number_format = pct_fmt

for c_idx, (sname, _) in enumerate(summary_p_cols, start=1):
    summary_p.column_dimensions[get_column_letter(c_idx)].width = max(14, min(40, len(sname) + 4))

priced_out = REPORTS / f"DyneFamilyUS_allTrusts_priced_{ts}.xlsx"
wb_p.save(priced_out)
print(f"Wrote: {priced_out.relative_to(PROJECT_ROOT)}")

# ----------------------------------------------------------------------------
# Console summary
# ----------------------------------------------------------------------------
print(f"\n=== Combined summary ===")
print(f"  trusts:           {len(trust_data)}")
print(f"  accounts:         {totals['# Accounts']}")
print(f"  rows:             {totals['# Rows']}")
print(f"  priced:           {total_priced_count}  "
      f"({total_priced_count / totals['# Rows'] * 100:.1f}%)")
print(f"  Masttro MV USD:   {total_masttro:>16,.2f}")
print(f"  Refreshed MV USD: {total_refreshed:>16,.2f}")
print(f"  Δ MV USD:         {total_delta:>16,.2f}  "
      f"({total_delta / total_masttro * 100:+.3f}%)")
