"""Export the tracker data into a single multi-sheet xlsx for browsing.

Pulls every API view for a given scope (defaults to Dyne Family US) and writes
each as a sheet. Useful for casual inspection without needing the Python REPL
or a SQLite browser.

Usage:
    python scripts/14_tracker_export_xlsx.py [scope_node_id]
"""

from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tracker import PROJECT_ROOT, api, compute
from tracker.api import connect

REPORTS = PROJECT_ROOT / "reports"

# Styles
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SECTION_FONT = Font(bold=True, italic=True, size=12)

MONEY_FMT = '#,##0.00;(#,##0.00);"-"'
PCT_FMT = '0.00%;-0.00%;"-"'
PRICE_FMT = '#,##0.0000'
INT_FMT = '#,##0'


def _detect_kind(col: str):
    """Heuristic for number formatting based on column name."""
    cl = col.lower()
    if cl in {"weight", "price_delta_pct", "period_return", "cumulative_twr",
              "dietz_return", "xirr", "top_1_weight", "top_5_weight", "top_10_weight"}:
        return "pct"
    if any(x in cl for x in ["mv_", "nav", "amount", "flow", "cost", "gain",
                              "income", "dividend", "interest", "valuation", "_usd"]):
        return "money"
    if "price" in cl:
        return "price"
    if cl in {"quantity", "n_positions", "n_accounts", "n_cashflows", "total_positions"}:
        return "qty"
    if cl == "hhi":
        return "qty"
    return "text"


def write_pivot(ws, df: pd.DataFrame, start_row: int = 1,
                title: str | None = None, value_kind: str = "pct",
                special_cols: dict | None = None) -> int:
    """Write a pivot DataFrame where every data column is the same numeric kind.

    special_cols: dict of {column_name: kind} to override for specific columns
    (e.g. a 'Total NAV' column should be money not pct).
    """
    special_cols = special_cols or {}
    r = start_row
    if title:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        r += 1
    if df is None or df.empty:
        ws.cell(row=r, column=1, value="(no data)")
        return r + 2
    if df.index.name:
        df = df.reset_index()
    cols = list(df.columns)
    # Header
    for c_idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=r, column=c_idx, value=str(name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[r].height = 32
    r += 1
    # Body
    for _, row in df.iterrows():
        for c_idx, name in enumerate(cols, start=1):
            v = row[name]
            if pd.isna(v):
                v = None
            cell = ws.cell(row=r, column=c_idx, value=v)
            if c_idx == 1:
                cell.font = Font(bold=True)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                kind = special_cols.get(name, value_kind if c_idx > 1 else "text")
                if kind == "money":
                    cell.number_format = MONEY_FMT
                elif kind == "pct":
                    cell.number_format = PCT_FMT
                elif kind == "qty":
                    cell.number_format = INT_FMT
                elif kind == "price":
                    cell.number_format = PRICE_FMT
        r += 1
    for c_idx, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(14, min(32, len(str(name)) + 4))
    return r + 1


def write_df(ws, df: pd.DataFrame, start_row: int = 1, title: str | None = None) -> int:
    """Write a DataFrame to a worksheet starting at start_row. Returns next free row."""
    r = start_row
    if title:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        r += 1
    if df is None or df.empty:
        ws.cell(row=r, column=1, value="(no data)")
        return r + 2

    # Reset index if it's a meaningful axis (date, period, etc.)
    if df.index.name and df.index.name not in ("",):
        df = df.reset_index()

    cols = list(df.columns)
    # Header
    for c_idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=r, column=c_idx, value=str(name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[r].height = 32
    r += 1

    # Body
    for _, row in df.iterrows():
        for c_idx, name in enumerate(cols, start=1):
            v = row[name]
            if isinstance(v, pd.Timestamp):
                v = v.strftime("%Y-%m-%d")
            elif pd.isna(v):
                v = None
            cell = ws.cell(row=r, column=c_idx, value=v)
            kind = _detect_kind(name)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if kind == "money":
                    cell.number_format = MONEY_FMT
                elif kind == "pct":
                    cell.number_format = PCT_FMT
                elif kind == "price":
                    cell.number_format = PRICE_FMT
                elif kind == "qty":
                    cell.number_format = INT_FMT
        r += 1

    # Column widths
    for c_idx, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(38, len(str(name)) + 4))
    return r + 1


def write_kv(ws, items: list[tuple[str, object]], start_row: int = 1,
             title: str | None = None) -> int:
    """Write key-value pairs vertically (for summary metrics)."""
    r = start_row
    if title:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        r += 1
    for label, value in items:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            kind = _detect_kind(label)
            if kind == "money":
                c.number_format = MONEY_FMT
            elif kind == "pct":
                c.number_format = PCT_FMT
            elif kind == "qty":
                c.number_format = INT_FMT
        r += 1
    return r + 1


def main():
    scope = sys.argv[1] if len(sys.argv) > 1 else "102_93356"

    conn = connect()
    cur = conn.cursor()
    cur.execute("SELECT alias, name FROM entity WHERE node_id = ?", (scope,))
    row = cur.fetchone()
    scope_label = (row[0] or row[1] or scope) if row else scope

    print(f"Exporting tracker view for: {scope_label} ({scope})")

    wb = Workbook()
    wb.remove(wb.active)

    # -------- Sheet 1: Summary --------
    ws = wb.create_sheet("Summary")
    perf = compute.performance_summary(conn, scope)
    conc = api.concentration(conn, scope)
    cp = api.current_positions(conn, scope, include_refresh=False)
    accts = api.scope_accounts(conn, scope)
    nav = api.nav_series(conn, scope)
    latest_nav = float(nav.iloc[-1]["nav"]) if not nav.empty else 0

    # Refresh stats
    cur.execute("SELECT MAX(refresh_date) FROM pricing_refresh")
    latest_refresh = cur.fetchone()[0]
    cp_priced = api.current_positions(conn, scope, include_refresh=True)
    refreshed_delta = (
        cp_priced["mv_refreshed_delta"].sum() if "mv_refreshed_delta" in cp_priced.columns else 0
    )

    headline = [
        ("Scope", scope_label),
        ("Scope nodeId", scope),
        ("Report generated", dt.datetime.now().isoformat(timespec="seconds")),
        ("Canonical accounts in scope", len(accts)),
        ("Distinct securities currently held", cp["security_id"].nunique() if not cp.empty else 0),
        ("Total positions", len(cp)),
        ("Current NAV (USD)", latest_nav),
        ("Latest yfinance refresh date", latest_refresh),
        ("Refresh Δ to NAV (USD)", float(refreshed_delta) if pd.notna(refreshed_delta) else 0),
    ]
    # 1-day return up in the headline
    one_day_h = api.one_day_return(conn, scope)
    if one_day_h and "error" not in one_day_h:
        headline.extend([
            ("1-day return as-of", one_day_h.get("as_of_date")),
            ("1-day previous close", one_day_h.get("previous_date")),
            ("1-day change (USD)", one_day_h.get("change_usd")),
            ("1-day return", one_day_h.get("return_pct")),
        ])
    r = write_kv(ws, headline, title="Headline metrics")

    if not perf.empty:
        ytd = perf[perf["period"] == "ytd"].iloc[0] if (perf["period"] == "ytd").any() else None
        one_y = perf[perf["period"] == "1y"].iloc[0] if (perf["period"] == "1y").any() else None
        if ytd is not None:
            r = write_kv(ws, [
                ("YTD return (Dietz)", float(ytd["dietz_return"]) if pd.notna(ytd["dietz_return"]) else None),
                ("YTD XIRR", float(ytd["xirr"]) if pd.notna(ytd["xirr"]) else None),
                ("YTD start NAV", float(ytd["start_nav"]) if pd.notna(ytd["start_nav"]) else None),
                ("YTD end NAV", float(ytd["end_nav"]) if pd.notna(ytd["end_nav"]) else None),
                ("YTD external flows", float(ytd["flows"]) if pd.notna(ytd["flows"]) else None),
                ("YTD gain (ex-flows)", float(ytd["gain"]) if pd.notna(ytd["gain"]) else None),
            ], start_row=r, title="Year-to-date")
        if one_y is not None:
            r = write_kv(ws, [
                ("1Y return (Dietz)", float(one_y["dietz_return"]) if pd.notna(one_y["dietz_return"]) else None),
                ("1Y XIRR", float(one_y["xirr"]) if pd.notna(one_y["xirr"]) else None),
            ], start_row=r, title="One-year")

    if conc:
        r = write_kv(ws, [(k, v) for k, v in conc.items()],
                     start_row=r, title="Concentration")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 32

    # -------- Sheet 2: Performance (aggregated, Masttro prices) --------
    ws = wb.create_sheet("Performance")
    r = write_df(ws, perf, title=f"Period performance — Masttro prices — {scope_label}")
    twr = compute.twr_series(conn, scope)
    write_df(ws, twr, start_row=r, title="TWR chain (per-period + cumulative) — Masttro prices")

    # -------- Sheet 2-refresh: Performance with yfinance refresh applied to end NAV --------
    ws = wb.create_sheet("Performance refreshed")
    r = write_df(ws, compute.performance_summary(conn, scope, use_refresh=True),
                 title=f"Period performance — yfinance-refreshed end NAV — {scope_label}")
    write_df(ws, compute.twr_series(conn, scope, use_refresh=True),
             start_row=r, title="TWR chain — yfinance-refreshed end NAV")

    # -------- Sheet 2-compare: side-by-side Masttro vs refreshed --------
    ws = wb.create_sheet("Performance comparison")
    write_df(ws, compute.performance_summary_comparison(conn, scope),
             title=f"Masttro vs yfinance-refreshed end NAV — {scope_label}")

    # -------- 1-day return section --------
    one_day = api.one_day_return(conn, scope)
    ws = wb.create_sheet("1-day return")
    if one_day and "error" not in one_day:
        r = write_kv(ws, [
            ("As-of (latest yfinance close)", one_day.get("as_of_date")),
            ("Previous close", one_day.get("previous_date")),
            ("NAV today (USD)", one_day.get("nav_today")),
            ("NAV yesterday (USD)", one_day.get("nav_yesterday")),
            ("Change (USD)", one_day.get("change_usd")),
            ("1-day return", one_day.get("return_pct")),
            ("Priced NAV (USD)", one_day.get("priced_nav")),
            ("Priced share of NAV", one_day.get("priced_pct")),
        ], title=f"1-day return — {scope_label}")
        r = write_df(ws, api.one_day_return_by(conn, scope, group_by="trust"),
                     start_row=r, title="1-day return by trust")
        write_df(ws, api.one_day_return_by(conn, scope, group_by="account"),
                 start_row=r, title="1-day return by account")
    else:
        ws.cell(row=1, column=1, value="(no 1-day return — pricing refresh missing previous-close data)")

    # -------- 1-day movers --------
    ws = wb.create_sheet("1-day movers")
    write_df(ws, api.one_day_movers(conn, scope, n=30),
             title="Top 30 contributors / detractors by USD impact (1-day)")

    # -------- Sheet 2b: Performance by trust --------
    ws = wb.create_sheet("Performance by trust")
    write_df(ws, compute.performance_by(conn, scope, group_by="trust"),
             title="Per-trust performance (one row per trust)")

    # -------- Sheet 2c: Performance by account --------
    ws = wb.create_sheet("Performance by account")
    write_df(ws, compute.performance_by(conn, scope, group_by="account"),
             title="Per-account performance (one row per account)")

    # -------- Sheet 3: NAV trajectory (aggregated) --------
    ws = wb.create_sheet("NAV")
    r = write_df(ws, nav, title=f"Monthly NAV — {scope_label}")

    # -------- Sheet 3b: NAV by trust --------
    ws_t = wb.create_sheet("NAV by trust")
    write_df(ws_t, api.nav_by(conn, scope, group_by="trust"),
             title="Monthly NAV by trust (columns = trusts, sorted by latest NAV)")

    # -------- Sheet 3c: NAV by account --------
    ws_a = wb.create_sheet("NAV by account")
    write_df(ws_a, api.nav_by(conn, scope, group_by="account"),
             title="Monthly NAV by account (columns = accounts, sorted by latest NAV)")

    # -------- Sheet 4: Current positions --------
    ws = wb.create_sheet("Current positions")
    write_df(ws, cp_priced, title=f"Latest snapshot: {cp['snapshot_date'].iloc[0] if not cp.empty else ''}")

    # -------- Sheet 5: Top 25 (aggregated) --------
    ws = wb.create_sheet("Top positions")
    write_df(ws, api.top_positions(conn, scope, n=25),
             title="Top 25 by market value (aggregated across accounts)")

    # -------- Sheet 5b: Top 10 by trust --------
    ws = wb.create_sheet("Top by trust")
    write_df(ws, api.top_positions_by(conn, scope, group_by="trust", n=10),
             title="Top 10 positions within each trust (long format)")

    # -------- Sheet 5c: Top 5 by account --------
    ws = wb.create_sheet("Top by account")
    write_df(ws, api.top_positions_by(conn, scope, group_by="account", n=5),
             title="Top 5 positions within each account (long format)")

    # -------- Sheet 6: Allocation (aggregated) --------
    ws = wb.create_sheet("Allocation")
    r = 1
    for by in ["asset_class", "security_type", "sector", "geographic_exposure",
               "custodian", "trust_alias"]:
        r = write_df(ws, api.allocation(conn, scope, by=by),
                     start_row=r, title=f"By {by}")

    # -------- Sheet 6b: Allocation by trust (pivots) --------
    ws = wb.create_sheet("Allocation by trust")
    r = 1
    for dim in ["asset_class", "security_type", "sector", "geographic_exposure", "custodian"]:
        df = api.allocation_by(conn, scope, group_by="trust", by_dimension=dim)
        r = write_pivot(ws, df, start_row=r,
                        title=f"Trust × {dim} (weights as % of trust NAV)",
                        special_cols={"Total NAV": "money"})

    # -------- Sheet 6c: Allocation by account (pivots) --------
    ws = wb.create_sheet("Allocation by account")
    r = 1
    for dim in ["asset_class", "security_type", "sector", "custodian"]:
        df = api.allocation_by(conn, scope, group_by="account", by_dimension=dim)
        r = write_pivot(ws, df, start_row=r,
                        title=f"Account × {dim} (weights as % of account NAV)",
                        special_cols={"Total NAV": "money"})

    # -------- Sheet 6d: Concentration by trust & account --------
    ws = wb.create_sheet("Concentration by trust")
    write_df(ws, api.concentration_by(conn, scope, group_by="trust"),
             title="Concentration metrics per trust")
    ws = wb.create_sheet("Concentration by account")
    write_df(ws, api.concentration_by(conn, scope, group_by="account"),
             title="Concentration metrics per account")

    # -------- Sheet 7: Income (aggregated by type) --------
    ws = wb.create_sheet("Income")
    write_df(ws, api.income_series(conn, scope, freq="M"),
             title="Monthly income (dividends + interest + other income)")

    # -------- Sheet 7b: Income by trust --------
    ws = wb.create_sheet("Income by trust")
    write_df(ws, api.income_by(conn, scope, group_by="trust", freq="M"),
             title="Monthly income by trust (columns = trusts)")

    # -------- Sheet 7c: Income by account --------
    ws = wb.create_sheet("Income by account")
    write_df(ws, api.income_by(conn, scope, group_by="account", freq="M"),
             title="Monthly income by account (columns = accounts)")

    # -------- Sheet 8: External flows (aggregated) --------
    ws = wb.create_sheet("External flows")
    r = write_df(ws, api.external_flows(conn, scope),
                 title="External capital flows (Deposits + Withdrawals) — every event")
    write_df(ws, api.external_flows_by(conn, scope, group_by="trust"),
             start_row=r, title="External flows totalled by trust")
    # Account-level totals are usually too many rows for the same sheet —
    # split into its own sheet.
    ws = wb.create_sheet("External flows by account")
    write_df(ws, api.external_flows_by(conn, scope, group_by="account"),
             title="External flows totalled by account")

    # -------- Sheet 9: Transactions --------
    ws = wb.create_sheet("Transactions")
    write_df(ws, api.transactions(conn, scope),
             title="All transactions (last 12 months)")

    # -------- Sheet 10: Trust roll-up --------
    ws = wb.create_sheet("Trusts")
    write_df(ws, api.trusts_in_scope(conn, scope),
             title=f"Trusts contributing to {scope_label}")

    # Save
    REPORTS.mkdir(exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    out = REPORTS / f"Tracker_{scope_label.replace(' ', '')}_{ts}.xlsx"
    wb.save(out)
    print(f"Wrote: {out.relative_to(PROJECT_ROOT)}")

    # Add a Sheet listing for convenience
    print(f"\nSheets in workbook:")
    for s in wb.sheetnames:
        print(f"  - {s}")


if __name__ == "__main__":
    main()
