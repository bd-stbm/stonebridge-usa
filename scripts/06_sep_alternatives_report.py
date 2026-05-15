"""Build an Excel report of all alternatives held by SEP Investments Pty Ltd.

Pure offline analysis - reads saved JSON in responses/, NO API calls.

Important structural facts (see memory project_subclient_mapping and the
diagnostic in this script):
- SEP Investments Pty Ltd is a SINGLE legal entity ($144,862,843.87 in
  positions as of 2026-05-13).
- It is beneficially owned by 5 Bermeister family members via 5 separate GWM
  pathways. Holdings rows under each SEP node carry fractional marketValues
  matching ownership share (Kevin 12.5%, Beverley 12.5%, Lisa 25%, Adam 25%,
  Nikki 25%).
- So the SAME alt appears as 5 rows (one per owner) with fractional figures
  in Holdings, Transactions and cef. This entity view sums the 5 slices
  back to the full SEP position.

Alt scope (broadened per user 2026-05-14):
- assetClass == 'Alternative Investments', OR
- has a cef row, OR
- has Capital Call / Return of Capital transactions
  (catches Fixed-Income-tagged corporate-debt feeders like Zagga, Arrowpoint).

Output: reports/SEP_alternatives_<YYYYMMDD-HHMMSS>.xlsx
"""

from __future__ import annotations

import datetime as dt
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESPONSES = PROJECT_ROOT / "responses"
REPORTS = PROJECT_ROOT / "reports"

SEP_NAME = "SEP Investments Pty Ltd"
ROOT_NODE = "0_7693"


def latest(pattern: str) -> Path:
    return sorted(RESPONSES.glob(pattern))[-1]


def load(pattern: str):
    p = latest(pattern)
    print(f"loaded: {p.relative_to(PROJECT_ROOT)}  ({p.stat().st_size:,} bytes)")
    return json.loads(p.read_text(encoding="utf-8"))


gwm = load("GWM-7693_*.json")
holdings = load("Holdings-7693_*.json")
transactions = load("Transactions-7693_*.json")
cef = load("cef-7693_*.json")

# -----------------------------------------------------------------------------
# Build GWM indices
# -----------------------------------------------------------------------------
nodes_by_id = {n["nodeId"]: n for n in gwm}
parent_of = {nid: n.get("parentNodeId") for nid, n in nodes_by_id.items()}
children_of: dict[str, list[str]] = defaultdict(list)
for nid, pid in parent_of.items():
    if pid:
        children_of[pid].append(nid)
direct_children_of_root = {nid for nid, pid in parent_of.items() if pid == ROOT_NODE}

sep_nodes = [n for n in gwm if n.get("alias") == SEP_NAME or n.get("name") == SEP_NAME]
sep_node_ids = {n["nodeId"] for n in sep_nodes}

# Subtree (union of all 5 ownership paths)
sep_subtree: set[str] = set(sep_node_ids)
stack = list(sep_node_ids)
while stack:
    nid = stack.pop()
    for c in children_of.get(nid, []):
        if c not in sep_subtree:
            sep_subtree.add(c)
            stack.append(c)

print(f"\nSEP entity: {len(sep_nodes)} GWM nodes (beneficial owners), "
      f"{len(sep_subtree)} total subtree nodes")


def under_sep(nid: str | None, pid: str | None) -> bool:
    cur = pid or nid
    for _ in range(60):
        if cur is None or cur == "_":
            return False
        if cur in sep_subtree:
            return True
        cur = parent_of.get(cur)
    return False


# -----------------------------------------------------------------------------
# Slice each dataset to SEP territory
# -----------------------------------------------------------------------------
sep_holdings = [h for h in holdings if under_sep(h.get("nodeId"), h.get("parentNodeId"))]
sep_txns = [t for t in transactions if under_sep(t.get("nodeId"), t.get("parentNodeId"))]
sep_cef = [c for c in cef if under_sep(c.get("nodeId"), c.get("parentNodeId"))]
print(f"  Holdings rows:    {len(sep_holdings):>5}")
print(f"  Transactions:     {len(sep_txns):>5}")
print(f"  cef rows:         {len(sep_cef):>5}")


# -----------------------------------------------------------------------------
# Asset grouping key — same fund across 5 owners should collapse to one group.
# Cef rows don't carry securityId, so we ONLY derive keys from Holdings here,
# and we'll attach cef + transactions to those keys by joining on nodeId.
# -----------------------------------------------------------------------------
def asset_key(row: dict) -> tuple:
    sid = row.get("securityId")
    if sid:
        return ("sid", sid)
    name = (row.get("assetName") or "").strip().lower()
    return ("name", name)


# Group Holdings by key, collect the set of nodeIds per group.
holdings_by_key: dict[tuple, list[dict]] = defaultdict(list)
nodeids_by_key: dict[tuple, set[str]] = defaultdict(set)
for h in sep_holdings:
    k = asset_key(h)
    holdings_by_key[k].append(h)
    nodeids_by_key[k].add(h["nodeId"])

# Reverse map: nodeId -> asset_key (so we can route cef and txn rows by nodeId).
key_by_nodeid: dict[str, tuple] = {}
for k, nids in nodeids_by_key.items():
    for nid in nids:
        key_by_nodeid[nid] = k

# Attach cef rows by nodeId.
cef_by_key: dict[tuple, list[dict]] = defaultdict(list)
cef_orphans: list[dict] = []
for c in sep_cef:
    k = key_by_nodeid.get(c.get("nodeId"))
    if k is None:
        cef_orphans.append(c)
    else:
        cef_by_key[k].append(c)

# Attach transactions by nodeId. Transactions on nodeIds not present in current
# Holdings (closed positions during the year) become orphans — capture them so
# YTD cashflows on fully-distributed funds aren't silently lost.
txns_by_key: dict[tuple, list[dict]] = defaultdict(list)
txn_orphans: list[dict] = []
for t in sep_txns:
    k = key_by_nodeid.get(t.get("nodeId"))
    if k is None:
        txn_orphans.append(t)
    else:
        txns_by_key[k].append(t)

print(f"\nJoined data into asset groups:")
print(f"  Holdings groups (distinct assets): {len(holdings_by_key)}")
print(f"  cef rows attached: {sum(len(v) for v in cef_by_key.values())}  "
      f"(orphans: {len(cef_orphans)})")
print(f"  txn rows attached: {sum(len(v) for v in txns_by_key.values())}  "
      f"(orphans: {len(txn_orphans)})")


# -----------------------------------------------------------------------------
# Decide which asset groups are "alt" — broadened scope
# -----------------------------------------------------------------------------
def has_alt_evidence(key: tuple) -> bool:
    if cef_by_key.get(key):
        return True
    for t in txns_by_key.get(key, []):
        if (t.get("transactionType") or "").strip() in {"Capital Call", "Return of Capital"}:
            return True
    return False


alt_keys = [
    k for k, rows in holdings_by_key.items()
    if any(r.get("assetClass") == "Alternative Investments" for r in rows)
    or has_alt_evidence(k)
]
print(f"\nDistinct alt assets under SEP (entity view): {len(alt_keys)}")


# -----------------------------------------------------------------------------
# Aggregate per alt
# -----------------------------------------------------------------------------
def first_non_null(rows: list[dict], field: str):
    for r in rows:
        v = r.get(field)
        if v not in (None, ""):
            return v
    return None


def sum_field(rows: list[dict], field: str) -> float:
    return sum((r.get(field) or 0) for r in rows)


def ytd_sum(txns: list[dict], ttype: str) -> float:
    s = 0.0
    for t in txns:
        if (t.get("transactionType") or "").strip() == ttype:
            s += t.get("netAmountRepCCY") or 0
    return s


rows_out = []
for key in alt_keys:
    h_rows = holdings_by_key.get(key, [])
    c_rows = cef_by_key.get(key, [])
    t_rows = txns_by_key.get(key, [])

    nav = sum_field(h_rows, "marketValue")
    commitment = sum_field(c_rows, "commitment")
    called = sum_field(c_rows, "capitalCalled")
    unfunded = sum_field(c_rows, "unfundedCommitment")
    distributed = sum_field(c_rows, "capitalDistributed")

    ytd_calls = -ytd_sum(t_rows, "Capital Call")        # flip outflow to positive
    ytd_roc = ytd_sum(t_rows, "Return of Capital")
    ytd_div = ytd_sum(t_rows, "Cash Dividends")
    ytd_income = ytd_sum(t_rows, "Income")
    ytd_interest = ytd_sum(t_rows, "Interest")
    ytd_mgmt = -ytd_sum(t_rows, "Management Fee")

    # Ratios / metadata from first non-null
    dpi = first_non_null(c_rows, "dpi")
    tvpi = first_non_null(c_rows, "tvpi")
    vintage = first_non_null(c_rows, "vintageYear")
    last_val = first_non_null(c_rows, "lastValuationDate")

    rows_out.append({
        "Asset Name": first_non_null(h_rows + c_rows, "assetName"),
        "Asset Class": first_non_null(h_rows, "assetClass"),
        "Security Type": first_non_null(h_rows + c_rows, "securityType"),
        "Local CCY": first_non_null(h_rows, "localCCY"),
        "# Owners": len(h_rows),
        "Commitment (AUD)": commitment if c_rows else None,
        "Capital Called - Cumulative (AUD)": called if c_rows else None,
        "Unfunded Commitment (AUD)": unfunded if c_rows else None,
        "Capital Distributed - Cumulative (AUD)": distributed if c_rows else None,
        "Current NAV (AUD)": nav,
        "YTD Capital Calls (AUD)": ytd_calls,
        "YTD Return of Capital (AUD)": ytd_roc,
        "YTD Cash Dividends (AUD)": ytd_div,
        "YTD Income (AUD)": ytd_income,
        "YTD Interest (AUD)": ytd_interest,
        "YTD Management Fee (AUD)": ytd_mgmt,
        "DPI": dpi,
        "TVPI": tvpi,
        "Vintage Year": vintage,
        "Last Valuation Date": (
            f"{last_val[:4]}-{last_val[4:6]}-{last_val[6:]}"
            if last_val and len(str(last_val)) == 8 else last_val
        ),
        "securityId": first_non_null(h_rows + c_rows, "securityId"),
    })

# Sort by NAV descending
rows_out.sort(key=lambda r: r["Current NAV (AUD)"] or 0, reverse=True)

# -----------------------------------------------------------------------------
# Write xlsx
# -----------------------------------------------------------------------------
COLUMNS = [
    ("Asset Name", "text"),
    ("Asset Class", "text"),
    ("Security Type", "text"),
    ("Local CCY", "text"),
    ("# Owners", "int"),
    ("Commitment (AUD)", "money"),
    ("Capital Called - Cumulative (AUD)", "money"),
    ("Unfunded Commitment (AUD)", "money"),
    ("Capital Distributed - Cumulative (AUD)", "money"),
    ("Current NAV (AUD)", "money"),
    ("YTD Capital Calls (AUD)", "money"),
    ("YTD Return of Capital (AUD)", "money"),
    ("YTD Cash Dividends (AUD)", "money"),
    ("YTD Income (AUD)", "money"),
    ("YTD Interest (AUD)", "money"),
    ("YTD Management Fee (AUD)", "money"),
    ("DPI", "ratio"),
    ("TVPI", "ratio"),
    ("Vintage Year", "text"),
    ("Last Valuation Date", "text"),
    ("securityId", "text"),
]

REPORTS.mkdir(exist_ok=True)
ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
out_path = REPORTS / f"SEP_alternatives_{ts}.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "SEP alternatives"

money_fmt = '#,##0.00;(#,##0.00);"-"'
ratio_fmt = '0.0000'

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
total_font = Font(bold=True)

for col_idx, (name, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

for r_idx, row in enumerate(rows_out, start=2):
    for c_idx, (name, kind) in enumerate(COLUMNS, start=1):
        v = row[name]
        cell = ws.cell(row=r_idx, column=c_idx, value=v)
        if kind == "money" and isinstance(v, (int, float)):
            cell.number_format = money_fmt
        elif kind == "ratio" and isinstance(v, (int, float)):
            cell.number_format = ratio_fmt
        elif kind == "int" and isinstance(v, (int, float)):
            cell.number_format = "0"

total_row_idx = len(rows_out) + 2
totals: dict[str, float] = {}
for name, kind in COLUMNS:
    if kind == "money":
        totals[name] = sum((r[name] or 0) for r in rows_out)

ws.cell(row=total_row_idx, column=1, value="TOTAL").font = total_font
for c_idx, (name, kind) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=total_row_idx, column=c_idx)
    cell.fill = total_fill
    cell.font = total_font
    if kind == "money":
        cell.value = totals[name]
        cell.number_format = money_fmt
    elif c_idx == 1:
        cell.value = "TOTAL"

ws.freeze_panes = "A2"
widths = {
    "Asset Name": 42,
    "Asset Class": 22,
    "Security Type": 28,
    "Local CCY": 10,
    "# Owners": 9,
    "Vintage Year": 12,
    "Last Valuation Date": 16,
    "securityId": 14,
    "DPI": 8,
    "TVPI": 8,
}
for c_idx, (name, _) in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(name, 22)
ws.row_dimensions[1].height = 38

wb.save(out_path)
print(f"\nWrote: {out_path.relative_to(PROJECT_ROOT)}")

# -----------------------------------------------------------------------------
# Console summary + sanity check
# -----------------------------------------------------------------------------
print(f"\n=== Entity-view totals ===")
print(f"  distinct alts:                  {len(rows_out)}")
print(f"  with commitment data (in cef):  {sum(1 for r in rows_out if r['Commitment (AUD)'])}")
print(f"  commitment-data-missing alts:   {sum(1 for r in rows_out if not r['Commitment (AUD)'])}")
print(f"  total commitment (AUD):         {totals['Commitment (AUD)']:>20,.2f}")
print(f"  total capital called (AUD):     {totals['Capital Called - Cumulative (AUD)']:>20,.2f}")
print(f"  total unfunded (AUD):           {totals['Unfunded Commitment (AUD)']:>20,.2f}")
print(f"  total distributed (AUD):        {totals['Capital Distributed - Cumulative (AUD)']:>20,.2f}")
print(f"  total current NAV (AUD):        {totals['Current NAV (AUD)']:>20,.2f}")
print(f"  YTD capital calls (AUD):        {totals['YTD Capital Calls (AUD)']:>20,.2f}")
print(f"  YTD return of capital (AUD):    {totals['YTD Return of Capital (AUD)']:>20,.2f}")
print(f"  YTD cash dividends (AUD):       {totals['YTD Cash Dividends (AUD)']:>20,.2f}")
print(f"  YTD income (AUD):               {totals['YTD Income (AUD)']:>20,.2f}")
print(f"  YTD interest (AUD):             {totals['YTD Interest (AUD)']:>20,.2f}")

# Sanity check: SEP total NAV should be < $144.86M (alts only is a subset)
all_h_nav = sum((h.get("marketValue") or 0) for h in sep_holdings)
print(f"\nSanity check:")
print(f"  Total SEP NAV (all assets, union of 5 owners):  {all_h_nav:>14,.2f}")
print(f"  Reported alt NAV:                                {totals['Current NAV (AUD)']:>14,.2f}")
print(f"  Alts as % of SEP total:                          {totals['Current NAV (AUD)']/all_h_nav*100:>13.1f}%")
