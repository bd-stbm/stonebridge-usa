"""Pull all directly-held holdings under any trust/entity's investment accounts.

Pure offline analysis — reads saved Holdings + GWM JSON, NO API calls.

Investment accounts = immediate children of the entity node with bankBroker
AND accountNumber populated. Holdings whose nodeId matches one of those
account nodeIds are "directly held" by the entity (vs. held via an SPV,
which are excluded).

Usage:
    python scripts/07_dylan_trust_holdings.py [trust_node_id]

Default trust_node_id is 102_93412 (Dylan Dyne Irrevocable Trust).
Output filename auto-derives from the trust's GWM alias.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESPONSES = PROJECT_ROOT / "responses"
REPORTS = PROJECT_ROOT / "reports"

TRUST_NODE_ID = sys.argv[1] if len(sys.argv) > 1 else "102_93412"


def _slug(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "", s or "")
    return s or "entity"


def latest(pattern: str) -> Path:
    return sorted(RESPONSES.glob(pattern))[-1]


def load(pattern: str):
    p = latest(pattern)
    print(f"loaded: {p.relative_to(PROJECT_ROOT)}  ({p.stat().st_size:,} bytes)")
    return json.loads(p.read_text(encoding="utf-8"))


gwm = load("GWM-7693_*.json")
holdings = load("Holdings-7693_*.json")

nodes_by_id = {n["nodeId"]: n for n in gwm}
children_of: dict[str, list[str]] = defaultdict(list)
for n in gwm:
    children_of[n.get("parentNodeId")].append(n["nodeId"])

if TRUST_NODE_ID not in nodes_by_id:
    raise SystemExit(f"Entity nodeId {TRUST_NODE_ID!r} not found in GWM.")
TRUST_NODE = nodes_by_id[TRUST_NODE_ID]
TRUST_NAME = TRUST_NODE.get("alias") or TRUST_NODE.get("name") or TRUST_NODE_ID

# Identify investment accounts: direct children of entity with bank+acct.
account_nodes = []
for child_nid in children_of.get(TRUST_NODE_ID, []):
    n = nodes_by_id[child_nid]
    bb = (n.get("bankBroker") or "").strip()
    an = (n.get("accountNumber") or "").strip()
    if bb and an:
        account_nodes.append(n)

print(f"\n{TRUST_NAME} (nodeId={TRUST_NODE_ID}) — "
      f"{len(account_nodes)} direct investment accounts:")
account_lookup: dict[str, dict] = {}
for n in account_nodes:
    nid = n["nodeId"]
    account_lookup[nid] = {
        "alias": n.get("alias"),
        "bank": n.get("bankBroker"),
        "account": n.get("accountNumber"),
        "valuation": n.get("valuation"),
    }
    print(f"  {nid:<12}  {str(n.get('bankBroker'))[:24]:<26} "
          f"{str(n.get('accountNumber'))[:18]:<20} {n.get('alias')}")

acct_nid_set = set(account_lookup.keys())

# Filter Holdings rows whose nodeId is one of the 7 accounts.
# Holdings.nodeId on a brokerage position = the account-level GWM node
# (Holdings.parentNodeId points one level above the account, i.e. the trust).
trust_holdings = [h for h in holdings if h.get("nodeId") in acct_nid_set]
print(f"\nHoldings rows directly under these {len(acct_nid_set)} accounts: {len(trust_holdings)}")

# Sort: by account valuation desc (so the biggest account comes first),
# then by marketValue desc within account.
acct_valuation = {nid: a["valuation"] or 0 for nid, a in account_lookup.items()}
trust_holdings.sort(
    key=lambda h: (
        -(acct_valuation.get(h["nodeId"], 0)),
        -(h.get("marketValue") or 0),
    )
)

# -----------------------------------------------------------------------------
# Build rows
# -----------------------------------------------------------------------------
COLUMNS = [
    ("Account Alias", "text"),
    ("Custodian", "text"),
    ("Account Number", "text"),
    ("Asset Name", "text"),
    ("ISIN", "text"),
    ("Ticker", "text"),
    ("SEDOL", "text"),
    ("CUSIP", "text"),
    ("Asset Class", "text"),
    ("Security Type", "text"),
    ("Sector", "text"),
    ("Geographic Exposure", "text"),
    ("Quantity", "qty"),
    ("Local CCY", "text"),
    ("Price (USD)", "price"),
    ("Market Value (USD)", "money"),
    ("Accrued Interest (USD)", "money"),
    ("Unit Cost (USD)", "price"),
    ("Total Cost (USD)", "money"),
    ("securityId", "text"),
    ("nodeId", "text"),
]

rows_out = []
for h in trust_holdings:
    acct = account_lookup[h["nodeId"]]
    qty = h.get("quantity")
    if isinstance(qty, str):
        try:
            qty = float(qty)
        except ValueError:
            pass
    # localAccruedInterest is in localCCY; Holdings.accruedInterest is in the
    # ccy= parameter we requested at pull time (AUD here). Since this trust is
    # 100% USD-denominated we use the local-currency fields, which ARE USD.
    rows_out.append({
        "Account Alias": acct["alias"],
        "Custodian": acct["bank"],
        "Account Number": acct["account"],
        "Asset Name": h.get("assetName"),
        "ISIN": h.get("isin"),
        "Ticker": h.get("ticker"),
        "SEDOL": h.get("sedol"),
        "CUSIP": h.get("cusip"),
        "Asset Class": h.get("assetClass"),
        "Security Type": h.get("securityType"),
        "Sector": h.get("sector"),
        "Geographic Exposure": h.get("geographicExposure"),
        "Quantity": qty,
        "Local CCY": h.get("localCCY"),
        "Price (USD)": h.get("price"),
        "Market Value (USD)": h.get("localMarketValue"),
        "Accrued Interest (USD)": h.get("localAccruedInterest"),
        "Unit Cost (USD)": h.get("unitCost"),
        "Total Cost (USD)": h.get("totalCost"),
        "securityId": h.get("securityId"),
        "nodeId": h.get("nodeId"),
    })

# -----------------------------------------------------------------------------
# Write xlsx
# -----------------------------------------------------------------------------
REPORTS.mkdir(exist_ok=True)
ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
out_path = REPORTS / f"{_slug(TRUST_NAME)}_holdings_{ts}.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Holdings"

money_fmt = '#,##0.00;(#,##0.00);"-"'
qty_fmt = '#,##0.0000;(#,##0.0000);"-"'
price_fmt = '#,##0.0000'

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
total_font = Font(bold=True)
subtotal_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
subtotal_font = Font(bold=True, italic=True)

# Header
for c_idx, (name, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=c_idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Write data with subtotal rows between accounts
r_idx = 2
current_acct = None
acct_start_row = 2
grand_total_mv_usd = 0.0
grand_total_accrued = 0.0
grand_total_cost = 0.0

def write_subtotal(start_row: int, end_row: int, label: str):
    """Emit a subtotal row using Excel SUM formulas for the numeric columns."""
    global r_idx
    if end_row < start_row:
        return
    money_cols = {
        "Market Value (USD)",
        "Accrued Interest (USD)", "Total Cost (USD)",
    }
    for c_idx, (name, kind) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = subtotal_fill
        cell.font = subtotal_font
        if c_idx == 1:
            cell.value = label
        elif name in money_cols:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            cell.number_format = money_fmt
    r_idx += 1

for row in rows_out:
    if current_acct is None:
        current_acct = row["Account Alias"]
        acct_start_row = r_idx
    elif row["Account Alias"] != current_acct:
        write_subtotal(acct_start_row, r_idx - 1, f"Subtotal: {current_acct}")
        current_acct = row["Account Alias"]
        acct_start_row = r_idx

    for c_idx, (name, kind) in enumerate(COLUMNS, start=1):
        v = row[name]
        cell = ws.cell(row=r_idx, column=c_idx, value=v)
        if kind == "money" and isinstance(v, (int, float)):
            cell.number_format = money_fmt
        elif kind == "qty" and isinstance(v, (int, float)):
            cell.number_format = qty_fmt
        elif kind == "price" and isinstance(v, (int, float)):
            cell.number_format = price_fmt

    if isinstance(row["Market Value (USD)"], (int, float)):
        grand_total_mv_usd += row["Market Value (USD)"]
    if isinstance(row["Accrued Interest (USD)"], (int, float)):
        grand_total_accrued += row["Accrued Interest (USD)"]
    if isinstance(row["Total Cost (USD)"], (int, float)):
        grand_total_cost += row["Total Cost (USD)"]
    r_idx += 1

# Last subtotal
if current_acct is not None:
    write_subtotal(acct_start_row, r_idx - 1, f"Subtotal: {current_acct}")

# Grand total
for c_idx, (name, kind) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=r_idx, column=c_idx)
    cell.fill = total_fill
    cell.font = total_font
    if c_idx == 1:
        cell.value = "GRAND TOTAL"
    elif name == "Market Value (USD)":
        cell.value = grand_total_mv_usd
        cell.number_format = money_fmt
    elif name == "Accrued Interest (USD)":
        cell.value = grand_total_accrued
        cell.number_format = money_fmt
    elif name == "Total Cost (USD)":
        cell.value = grand_total_cost
        cell.number_format = money_fmt

# Column widths
widths = {
    "Account Alias": 28,
    "Custodian": 18,
    "Account Number": 14,
    "Asset Name": 38,
    "ISIN": 14,
    "Ticker": 10,
    "SEDOL": 10,
    "CUSIP": 12,
    "Asset Class": 22,
    "Security Type": 22,
    "Sector": 22,
    "Geographic Exposure": 20,
    "Quantity": 14,
    "Local CCY": 8,
    "Price (USD)": 12,
    "Market Value (USD)": 18,
    "Accrued Interest (USD)": 16,
    "Unit Cost (USD)": 12,
    "Total Cost (USD)": 18,
    "securityId": 12,
    "nodeId": 12,
}
for c_idx, (name, _) in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(name, 14)

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 38
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

wb.save(out_path)
print(f"\nWrote: {out_path.relative_to(PROJECT_ROOT)}")

# -----------------------------------------------------------------------------
# Console summary + sanity check
# -----------------------------------------------------------------------------
print(f"\n=== Summary ===")
print(f"  rows:                            {len(rows_out)}")
print(f"  grand total market value (USD):  {grand_total_mv_usd:>16,.2f}")

# Per-account summary
print(f"\nPer-account row count + USD MV sum:")
by_acct = defaultdict(lambda: [0, 0.0])  # rows, mv
for r in rows_out:
    by_acct[r["Account Alias"]][0] += 1
    if isinstance(r["Market Value (USD)"], (int, float)):
        by_acct[r["Account Alias"]][1] += r["Market Value (USD)"]
for alias, (rc, mv) in sorted(by_acct.items(), key=lambda kv: -kv[1][1]):
    print(f"  {alias[:38]:<40}  rows={rc:>4}  MV(USD)={mv:>14,.2f}")

# Asset class breakdown
from collections import Counter
ac = Counter(r["Asset Class"] for r in rows_out)
print(f"\nBy assetClass:")
for cls, n in ac.most_common():
    print(f"  {str(cls):<28}  {n}")

# ISIN coverage
isin_filled = sum(1 for r in rows_out if (r["ISIN"] or "").strip())
print(f"\nISIN coverage: {isin_filled}/{len(rows_out)}")
