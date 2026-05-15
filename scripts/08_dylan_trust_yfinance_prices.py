"""Enrich a trust's/entity's holdings with yfinance pricing (+ OpenFIGI fallback).

Reads the saved Masttro Holdings JSON, filters to the directly-held investment
accounts of the given entity, normalises tickers, batches a download from
yfinance, falls back to OpenFIGI ISIN→ticker resolution for failures, and
writes a new xlsx alongside Masttro prices for comparison.

Ticker normalisation:
- uppercase
- strip Masttro country suffixes (_US, _AU, _LN, _GB, _FR, _DE, _JP, _CA, _HK)
- replace "/" with "-" (Berkshire B-shares etc.)

Usage:
    python scripts/08_dylan_trust_yfinance_prices.py [trust_node_id]

Default trust_node_id is 102_93412 (Dylan Dyne Irrevocable Trust).
Output filename auto-derives from the entity's GWM alias.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
import urllib.error
import urllib.request
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
# Windows console defaults to cp1252; force UTF-8 so Δ etc. print cleanly.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

OPENFIGI_URL = "https://api.openfigi.com/v3/mapping"
# Order of preference for US-listed exchange codes when picking a ticker
# from OpenFIGI's match set. "US" = composite (preferred for yfinance).
US_EXCH_PRIORITY = ["US", "UN", "UQ", "UA", "UR", "UF", "UV", "UB", "UW", "UP"]

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESPONSES = PROJECT_ROOT / "responses"
REPORTS = PROJECT_ROOT / "reports"

TRUST_NODE_ID = sys.argv[1] if len(sys.argv) > 1 else "102_93412"
COUNTRY_SUFFIX_RE = re.compile(r"_(US|AU|LN|GB|FR|DE|JP|CA|HK)$", re.IGNORECASE)


def _slug(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "", s or "")
    return s or "entity"


def normalize_ticker(t: str | None) -> str | None:
    if not t:
        return None
    t = t.strip().upper()
    t = COUNTRY_SUFFIX_RE.sub("", t)
    # yfinance expects "-" not "/" for share-class delimiters (BRK/B -> BRK-B).
    t = t.replace("/", "-")
    return t or None


def openfigi_resolve(isins: list[str], chunk: int = 10) -> dict[str, str | None]:
    """Batch-resolve ISINs to a US-listed ticker via OpenFIGI /v3/mapping.

    No API key needed for low-volume calls (25 req/6s, max 10 mappings per req).
    Returns {isin: ticker | None}. Preference order:
      1. exchCode == 'US' (US composite — the right one for yfinance)
      2. Any US exchange code (UN/UQ/UA/...)
      3. First match with a non-null ticker
    """
    out: dict[str, str | None] = {}
    if not isins:
        return out
    for i in range(0, len(isins), chunk):
        batch = isins[i:i + chunk]
        body = json.dumps(
            [{"idType": "ID_ISIN", "idValue": v} for v in batch]
        ).encode("utf-8")
        req = urllib.request.Request(
            OPENFIGI_URL,
            data=body,
            headers={"Content-Type": "application/json", "Accept": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                response = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")[:200]
            print(f"  OpenFIGI batch error: HTTP {e.code} {err_body}")
            for v in batch:
                out[v] = None
            continue
        except Exception as e:
            print(f"  OpenFIGI batch error: {type(e).__name__}: {e}")
            for v in batch:
                out[v] = None
            continue
        for isin_value, result in zip(batch, response):
            matches = result.get("data") or []
            chosen = None
            for ex in US_EXCH_PRIORITY:
                chosen = next(
                    (m for m in matches if m.get("exchCode") == ex and m.get("ticker")),
                    None,
                )
                if chosen:
                    break
            if not chosen:
                chosen = next((m for m in matches if m.get("ticker")), None)
            out[isin_value] = chosen["ticker"] if chosen else None
    return out


def latest(pattern: str) -> Path:
    return sorted(RESPONSES.glob(pattern))[-1]


def load(pattern: str):
    p = latest(pattern)
    print(f"loaded: {p.relative_to(PROJECT_ROOT)}  ({p.stat().st_size:,} bytes)")
    return json.loads(p.read_text(encoding="utf-8"))


# -----------------------------------------------------------------------------
# Load + filter
# -----------------------------------------------------------------------------
gwm = load("GWM-7693_*.json")
holdings = load("Holdings-7693_*.json")

nodes_by_id = {n["nodeId"]: n for n in gwm}
children_of: dict[str, list[str]] = defaultdict(list)
for n in gwm:
    children_of[n.get("parentNodeId")].append(n["nodeId"])

if TRUST_NODE_ID not in nodes_by_id:
    raise SystemExit(f"Entity nodeId {TRUST_NODE_ID!r} not found in GWM.")
TRUST_NODE = nodes_by_id[TRUST_NODE_ID]
TRUST_NAME = TRUST_NODE.get("alias") or TRUST_NODE.get("name") or TRUST_NODE_ID

account_nodes = []
for child_nid in children_of.get(TRUST_NODE_ID, []):
    n = nodes_by_id[child_nid]
    if (n.get("bankBroker") or "").strip() and (n.get("accountNumber") or "").strip():
        account_nodes.append(n)
acct_nid_set = {n["nodeId"] for n in account_nodes}
account_lookup = {n["nodeId"]: n for n in account_nodes}

trust_holdings = [h for h in holdings if h.get("nodeId") in acct_nid_set]
print(f"\n{TRUST_NAME} (nodeId={TRUST_NODE_ID}): "
      f"{len(trust_holdings)} holdings across {len(acct_nid_set)} accounts")


# -----------------------------------------------------------------------------
# Build the unique normalised-ticker set we need to price
# -----------------------------------------------------------------------------
ticker_map: dict[int, str | None] = {}  # row index -> normalised ticker
unique_tickers: set[str] = set()
# Restrict to Equity only. Individual bonds (Fixed Income) and Commodities
# return false-positive matches from yfinance on CUSIP-derived "tickers"
# — e.g. a muni-bond row priced as some unrelated equity. ETFs and bond ETFs
# are tagged 'Equity' by Masttro, so this filter keeps them.
for i, h in enumerate(trust_holdings):
    nt = normalize_ticker(h.get("ticker"))
    if nt and h.get("assetClass") == "Equity":
        ticker_map[i] = nt
        unique_tickers.add(nt)

print(f"Unique normalised tickers to look up: {len(unique_tickers)}")


# -----------------------------------------------------------------------------
# Batch-download from yfinance
# -----------------------------------------------------------------------------
def fetch_prices(tickers: list[str], chunk: int = 50) -> tuple[dict, list]:
    """Returns (ok={ticker: (price, asof_date, currency_or_None)}, failed=[tickers])"""
    ok: dict[str, tuple] = {}
    failed: list[str] = []
    for i in range(0, len(tickers), chunk):
        batch = tickers[i:i + chunk]
        print(f"  batch {i//chunk + 1}: {len(batch)} tickers...", end=" ", flush=True)
        try:
            data = yf.download(
                tickers=batch,
                period="5d",
                group_by="ticker",
                auto_adjust=False,
                progress=False,
                threads=True,
            )
        except Exception as e:
            print(f"ERR: {e}")
            failed.extend(batch)
            continue
        got = 0
        for t in batch:
            try:
                if len(batch) == 1:
                    closes = data["Close"].dropna()
                else:
                    # multi-ticker shape: data has top-level MultiIndex columns
                    closes = data[t]["Close"].dropna()
                if len(closes):
                    ok[t] = (float(closes.iloc[-1]), closes.index[-1].date(), None)
                    got += 1
                else:
                    failed.append(t)
            except (KeyError, TypeError):
                failed.append(t)
        print(f"got {got}/{len(batch)}")
    return ok, failed


print(f"\nFetching prices from yfinance (pass 1: Masttro tickers)...")
prices, failures = fetch_prices(sorted(unique_tickers))
print(f"  Priced: {len(prices)}/{len(unique_tickers)}   Failed: {len(failures)}")

# ----------------------------------------------------------------------
# OpenFIGI fallback: for tickers that failed, look up by ISIN.
# ----------------------------------------------------------------------
# Track which ticker each row uses and where it came from.
# ticker_source[i] in {"masttro", "openfigi", None}
ticker_source: dict[int, str] = {}
for i, t in ticker_map.items():
    if t in prices:
        ticker_source[i] = "masttro"

# Collect ISINs for the failed tickers (one per ticker is enough).
failed_set = set(failures)
isin_by_failed_ticker: dict[str, str] = {}
for i, t in ticker_map.items():
    if t in failed_set and t not in isin_by_failed_ticker:
        isin = (trust_holdings[i].get("isin") or "").strip()
        if isin:
            isin_by_failed_ticker[t] = isin

isins_to_resolve = sorted(set(isin_by_failed_ticker.values()))
print(f"\nOpenFIGI fallback: resolving {len(isins_to_resolve)} ISINs "
      f"(from {len(failed_set)} failed tickers, "
      f"{len(failed_set) - len(isin_by_failed_ticker)} have no ISIN)...")
isin_to_ticker = openfigi_resolve(isins_to_resolve)
resolved_count = sum(1 for v in isin_to_ticker.values() if v)
print(f"  Resolved: {resolved_count}/{len(isins_to_resolve)}")

# Build remap: old ticker -> new ticker (via OpenFIGI)
ticker_remap: dict[str, str] = {}
for old_ticker, isin in isin_by_failed_ticker.items():
    new = isin_to_ticker.get(isin)
    if new and new != old_ticker:
        # Normalise yfinance-style (handles BRK/B -> BRK-B if it comes back that way)
        new_norm = normalize_ticker(new)
        if new_norm:
            ticker_remap[old_ticker] = new_norm

new_unique = sorted(set(ticker_remap.values()) - set(prices.keys()))
print(f"  New unique tickers from OpenFIGI to try: {len(new_unique)}")

if new_unique:
    print(f"\nFetching prices from yfinance (pass 2: OpenFIGI tickers)...")
    extra_prices, extra_failures = fetch_prices(new_unique)
    print(f"  Priced: {len(extra_prices)}/{len(new_unique)}   "
          f"Failed: {len(extra_failures)}")
    prices.update(extra_prices)

# Apply the remap to ticker_map for any row whose original ticker was in failed_set
for i, t in list(ticker_map.items()):
    if t in ticker_remap:
        new = ticker_remap[t]
        if new in prices:
            ticker_map[i] = new
            ticker_source[i] = "openfigi"

# Build the final failure list (tickers still unresolved)
still_failed = sorted(
    {ticker_map[i] for i in ticker_map}
    - set(prices.keys())
)
print(f"\nFinal: priced={len(set(ticker_map.values()) & set(prices.keys()))}, "
      f"still unresolved={len(still_failed)}")
if still_failed[:20]:
    print(f"  Still failed: {still_failed[:20]}")


# -----------------------------------------------------------------------------
# Build rows for xlsx
# -----------------------------------------------------------------------------
COLUMNS = [
    ("Account Alias", "text"),
    ("Custodian", "text"),
    ("Account Number", "text"),
    ("Asset Name", "text"),
    ("ISIN", "text"),
    ("Ticker (Masttro)", "text"),
    ("Ticker (yfinance)", "text"),
    ("Ticker Source", "text"),
    ("SEDOL", "text"),
    ("CUSIP", "text"),
    ("Asset Class", "text"),
    ("Security Type", "text"),
    ("Sector", "text"),
    ("Geographic Exposure", "text"),
    ("Quantity", "qty"),
    ("Local CCY", "text"),
    ("Masttro Price (USD)", "price"),
    ("yfinance Price (USD)", "price"),
    ("yfinance As-of", "text"),
    ("Δ Price (USD)", "price"),
    ("Δ Price %", "pct"),
    ("Masttro MV (USD)", "money"),
    ("Refreshed MV (USD)", "money"),
    ("Refreshed Δ MV (USD)", "money"),
    ("Accrued Interest (USD)", "money"),
    ("Unit Cost (USD)", "price"),
    ("Total Cost (USD)", "money"),
    ("securityId", "text"),
    ("nodeId", "text"),
]

# Sort by account valuation desc, then marketValue desc
acct_val = {nid: n.get("valuation") or 0 for nid, n in account_lookup.items()}
trust_holdings_sorted = sorted(
    range(len(trust_holdings)),
    key=lambda i: (
        -(acct_val.get(trust_holdings[i]["nodeId"], 0)),
        -(trust_holdings[i].get("marketValue") or 0),
    ),
)

rows_out = []
unpriced_count = 0
for idx in trust_holdings_sorted:
    h = trust_holdings[idx]
    acct = account_lookup[h["nodeId"]]
    yf_ticker = ticker_map.get(idx)
    yf_price = None
    yf_date = None
    yf_source = None
    if yf_ticker and yf_ticker in prices:
        yf_price, yf_date, _ = prices[yf_ticker]
        yf_source = ticker_source.get(idx, "masttro")
    elif yf_ticker:
        unpriced_count += 1

    masttro_price = h.get("price")
    qty = h.get("quantity")
    if isinstance(qty, str):
        try:
            qty = float(qty)
        except ValueError:
            qty = None

    delta_price = None
    delta_pct = None
    refreshed_usd_mv = None
    delta_usd_mv = None
    masttro_usd_mv = h.get("localMarketValue")  # USD-native for this trust
    if (
        isinstance(masttro_price, (int, float))
        and isinstance(yf_price, (int, float))
        and masttro_price != 0
    ):
        delta_price = yf_price - masttro_price
        delta_pct = delta_price / masttro_price
        # Quantity is None on many Masttro ETF rows, so derive Refreshed MV
        # from the price ratio applied to Masttro's existing local MV.
        # Both Masttro and yfinance prices are in USD, so no FX involved.
        if isinstance(masttro_usd_mv, (int, float)) and masttro_usd_mv != 0:
            refreshed_usd_mv = masttro_usd_mv * (yf_price / masttro_price)
            delta_usd_mv = refreshed_usd_mv - masttro_usd_mv

    rows_out.append({
        "Account Alias": acct.get("alias"),
        "Custodian": acct.get("bankBroker"),
        "Account Number": acct.get("accountNumber"),
        "Asset Name": h.get("assetName"),
        "ISIN": h.get("isin"),
        "Ticker (Masttro)": h.get("ticker"),
        "Ticker (yfinance)": yf_ticker if yf_price is not None else None,
        "Ticker Source": yf_source,
        "SEDOL": h.get("sedol"),
        "CUSIP": h.get("cusip"),
        "Asset Class": h.get("assetClass"),
        "Security Type": h.get("securityType"),
        "Sector": h.get("sector"),
        "Geographic Exposure": h.get("geographicExposure"),
        "Quantity": qty,
        "Local CCY": h.get("localCCY"),
        "Masttro Price (USD)": masttro_price,
        "yfinance Price (USD)": yf_price,
        "yfinance As-of": yf_date.isoformat() if yf_date else None,
        "Δ Price (USD)": delta_price,
        "Δ Price %": delta_pct,
        "Masttro MV (USD)": masttro_usd_mv,
        "Refreshed MV (USD)": refreshed_usd_mv,
        "Refreshed Δ MV (USD)": delta_usd_mv,
        "Accrued Interest (USD)": h.get("localAccruedInterest"),
        "Unit Cost (USD)": h.get("unitCost"),
        "Total Cost (USD)": h.get("totalCost"),
        "securityId": h.get("securityId"),
        "nodeId": h.get("nodeId"),
    })

# -----------------------------------------------------------------------------
# Write xlsx
# -----------------------------------------------------------------------------
REPORTS.mkdir(exist_ok=True)
ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
out_path = REPORTS / f"{_slug(TRUST_NAME)}_priced_{ts}.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Holdings (priced)"

money_fmt = '#,##0.00;(#,##0.00);"-"'
qty_fmt = '#,##0.0000;(#,##0.0000);"-"'
price_fmt = '#,##0.0000'
pct_fmt = '0.00%;-0.00%;"-"'

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
total_font = Font(bold=True)
subtotal_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
subtotal_font = Font(bold=True, italic=True)

for c_idx, (name, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=c_idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

money_cols_for_subtotal = {
    "Masttro MV (USD)", "Refreshed MV (USD)", "Refreshed Δ MV (USD)",
    "Accrued Interest (USD)", "Total Cost (USD)",
}

r_idx = 2
current_acct = None
acct_start = 2

def write_subtotal(start_row: int, end_row: int, label: str):
    global r_idx
    if end_row < start_row:
        return
    for c_idx, (name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = subtotal_fill
        cell.font = subtotal_font
        if c_idx == 1:
            cell.value = label
        elif name in money_cols_for_subtotal:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            cell.number_format = money_fmt
    r_idx += 1

for row in rows_out:
    if current_acct is None:
        current_acct = row["Account Alias"]
        acct_start = r_idx
    elif row["Account Alias"] != current_acct:
        write_subtotal(acct_start, r_idx - 1, f"Subtotal: {current_acct}")
        current_acct = row["Account Alias"]
        acct_start = r_idx

    for c_idx, (name, kind) in enumerate(COLUMNS, start=1):
        v = row[name]
        cell = ws.cell(row=r_idx, column=c_idx, value=v)
        if kind == "money" and isinstance(v, (int, float)):
            cell.number_format = money_fmt
        elif kind == "qty" and isinstance(v, (int, float)):
            cell.number_format = qty_fmt
        elif kind == "price" and isinstance(v, (int, float)):
            cell.number_format = price_fmt
        elif kind == "pct" and isinstance(v, (int, float)):
            cell.number_format = pct_fmt
    r_idx += 1

if current_acct is not None:
    write_subtotal(acct_start, r_idx - 1, f"Subtotal: {current_acct}")

# Grand total
for c_idx, (name, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=r_idx, column=c_idx)
    cell.fill = total_fill
    cell.font = total_font
    if c_idx == 1:
        cell.value = "GRAND TOTAL"
    elif name in money_cols_for_subtotal:
        col_letter = get_column_letter(c_idx)
        cell.value = (
            f"=SUMIF($A$2:$A${r_idx-1},\"<>Subtotal*\","
            f"{col_letter}$2:{col_letter}${r_idx-1})"
        )
        cell.number_format = money_fmt

widths = {
    "Account Alias": 28, "Custodian": 18, "Account Number": 14,
    "Asset Name": 38, "ISIN": 14, "Ticker (Masttro)": 14, "Ticker (yfinance)": 14, "Ticker Source": 11,
    "SEDOL": 10, "CUSIP": 12, "Asset Class": 22, "Security Type": 22,
    "Sector": 22, "Geographic Exposure": 20, "Quantity": 14, "Local CCY": 8,
    "Masttro Price (USD)": 14, "yfinance Price (USD)": 14,
    "yfinance As-of": 12, "Δ Price (USD)": 12, "Δ Price %": 10,
    "Masttro MV (USD)": 18, "Refreshed MV (USD)": 18, "Refreshed Δ MV (USD)": 16,
    "Accrued Interest (USD)": 14, "Unit Cost (USD)": 12,
    "Total Cost (USD)": 16, "securityId": 12, "nodeId": 12,
}
for c_idx, (name, _) in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(name, 14)

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 38
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

wb.save(out_path)
print(f"\nWrote: {out_path.relative_to(PROJECT_ROOT)}")

# -----------------------------------------------------------------------------
# Summary
# -----------------------------------------------------------------------------
priced = sum(1 for r in rows_out if isinstance(r["yfinance Price (USD)"], (int, float)))
total = len(rows_out)
priced_via_masttro = sum(1 for r in rows_out if r["Ticker Source"] == "masttro")
priced_via_openfigi = sum(1 for r in rows_out if r["Ticker Source"] == "openfigi")
not_eligible = sum(
    1 for r in rows_out
    if not normalize_ticker(r["Ticker (Masttro)"]) or r["Asset Class"] != "Equity"
)
eligible_unpriced = total - priced - not_eligible

sum_refreshed_delta = sum(
    (r["Refreshed Δ MV (USD)"] or 0) for r in rows_out
)
sum_masttro_usd = sum(
    (r["Masttro MV (USD)"] or 0) for r in rows_out
)
sum_refreshed_usd = sum(
    (r["Refreshed MV (USD)"] or 0) for r in rows_out
    if isinstance(r["Refreshed MV (USD)"], (int, float))
)

print(f"\n=== Pricing summary ===")
print(f"  rows total:                  {total}")
print(f"  priced via yfinance:         {priced}")
print(f"    via Masttro ticker:        {priced_via_masttro}")
print(f"    via OpenFIGI fallback:     {priced_via_openfigi}")
print(f"  not eligible for lookup:     {not_eligible}  (cash, bonds w/o ticker, etc.)")
print(f"  eligible but unpriced:       {eligible_unpriced}  (delisted, OTC, name change, etc.)")
print(f"\n  Masttro MV (USD):            {sum_masttro_usd:>16,.2f}")
print(f"  Refreshed MV (USD, priced rows only): {sum_refreshed_usd:>14,.2f}")
print(f"  Δ on priced rows (USD):      {sum_refreshed_delta:>16,.2f}  "
      f"({sum_refreshed_delta/sum_masttro_usd*100:+.3f}% of total Masttro MV)")

print(f"\n  Top 10 absolute price movers (by USD delta on the position):")
movers = [r for r in rows_out if isinstance(r["Refreshed Δ MV (USD)"], (int, float))]
movers.sort(key=lambda r: abs(r["Refreshed Δ MV (USD)"]), reverse=True)
for r in movers[:10]:
    print(f"    {str(r['Asset Name'])[:38]:<40} "
          f"{r['Ticker (yfinance)']:<6} "
          f"Δprice={r['Δ Price %']*100:+.2f}%  "
          f"Δ USD MV={r['Refreshed Δ MV (USD)']:>14,.2f}")
